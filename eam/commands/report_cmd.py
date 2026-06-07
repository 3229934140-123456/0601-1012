import csv
import click
from tabulate import tabulate
from ..database import (
    get_db, query_assets, get_operation_logs, get_asset_statistics,
    get_asset_by_no, ASSET_CATEGORIES, ASSET_STATUS
)


@click.command('report')
@click.option('--type', '-t', 'report_type',
              type=click.Choice(['summary', 'by-category', 'by-department', 'by-status',
                                 'depreciation', 'value']),
              default='summary', help='报表类型')
@click.option('--export', '-e', help='导出到文件')
@click.option('--format', '-f', 'fmt', type=click.Choice(['csv', 'txt']), default='txt',
              help='导出格式')
def report_cmd(report_type, export, fmt):
    """生成资产报表"""
    with get_db() as conn:
        if report_type == 'summary':
            _report_summary(conn, export, fmt)
        elif report_type == 'by-category':
            _report_by_category(conn, export, fmt)
        elif report_type == 'by-department':
            _report_by_department(conn, export, fmt)
        elif report_type == 'by-status':
            _report_by_status(conn, export, fmt)
        elif report_type == 'depreciation':
            _report_depreciation(conn, export, fmt)
        elif report_type == 'value':
            _report_value(conn, export, fmt)


def _report_summary(conn, export, fmt):
    stats = get_asset_statistics(conn)

    click.echo("=" * 50)
    click.echo("  资产总览报表")
    click.echo("=" * 50)
    click.echo(f"  资产总数: {stats['total']} 件")
    click.echo(f"  资产总值: {stats['total_value']:.2f} 元")
    click.echo()
    click.echo("  按类别分布:")
    for cat, count in stats['by_category'].items():
        click.echo(f"    {cat}: {count} 件")
    click.echo()
    click.echo("  按状态分布:")
    for status, count in stats['by_status'].items():
        click.echo(f"    {status}: {count} 件")
    click.echo()
    click.echo("  按部门分布:")
    for dept, count in stats['by_department'].items():
        click.echo(f"    {dept or '未分配'}: {count} 件")
    click.echo("=" * 50)

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['指标', '数值'])
            writer.writerow(['资产总数', stats['total']])
            writer.writerow(['资产总值', f"{stats['total_value']:.2f}"])
            writer.writerow([])
            writer.writerow(['按类别分布'])
            for cat, count in stats['by_category'].items():
                writer.writerow([cat, count])
            writer.writerow([])
            writer.writerow(['按状态分布'])
            for status, count in stats['by_status'].items():
                writer.writerow([status, count])
            writer.writerow([])
            writer.writerow(['按部门分布'])
            for dept, count in stats['by_department'].items():
                writer.writerow([dept or '未分配', count])
        click.echo(f"已导出到 {export}")


def _report_by_category(conn, export, fmt):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT category, status, COUNT(*) as count, SUM(purchase_price) as total_value
        FROM assets
        GROUP BY category, status
        ORDER BY category, status
    ''')
    rows = cursor.fetchall()

    categories = {}
    for row in rows:
        cat = row['category']
        if cat not in categories:
            categories[cat] = {'总数': 0, '总价值': 0, '状态': {}}
        categories[cat]['总数'] += row['count']
        categories[cat]['总价值'] += row['total_value'] or 0
        categories[cat]['状态'][row['status']] = row['count']

    table = []
    for cat, data in categories.items():
        status_str = ', '.join(f"{k}:{v}" for k, v in data['状态'].items())
        table.append([cat, data['总数'], f"{data['总价值']:.2f}", status_str])

    headers = ['类别', '数量', '总价值(元)', '状态分布']
    click.echo(tabulate(table, headers=headers, tablefmt='simple'))

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(table)
        click.echo(f"已导出到 {export}")


def _report_by_department(conn, export, fmt):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT department, category, COUNT(*) as count, SUM(purchase_price) as total_value
        FROM assets
        GROUP BY department, category
        ORDER BY department, category
    ''')
    rows = cursor.fetchall()

    departments = {}
    for row in rows:
        dept = row['department'] or '未分配'
        if dept not in departments:
            departments[dept] = {'总数': 0, '总价值': 0, '类别': {}}
        departments[dept]['总数'] += row['count']
        departments[dept]['总价值'] += row['total_value'] or 0
        departments[dept]['类别'][row['category']] = row['count']

    table = []
    for dept, data in departments.items():
        cat_str = ', '.join(f"{k}:{v}" for k, v in data['类别'].items())
        table.append([dept, data['总数'], f"{data['总价值']:.2f}", cat_str])

    headers = ['部门', '数量', '总价值(元)', '类别分布']
    click.echo(tabulate(table, headers=headers, tablefmt='simple'))

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(table)
        click.echo(f"已导出到 {export}")


def _report_by_status(conn, export, fmt):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT status, category, COUNT(*) as count, SUM(purchase_price) as total_value
        FROM assets
        GROUP BY status, category
        ORDER BY status, category
    ''')
    rows = cursor.fetchall()

    statuses = {}
    for row in rows:
        status = row['status']
        if status not in statuses:
            statuses[status] = {'总数': 0, '总价值': 0, '类别': {}}
        statuses[status]['总数'] += row['count']
        statuses[status]['总价值'] += row['total_value'] or 0
        statuses[status]['类别'][row['category']] = row['count']

    table = []
    for status, data in statuses.items():
        cat_str = ', '.join(f"{k}:{v}" for k, v in data['类别'].items())
        table.append([status, data['总数'], f"{data['总价值']:.2f}", cat_str])

    headers = ['状态', '数量', '总价值(元)', '类别分布']
    click.echo(tabulate(table, headers=headers, tablefmt='simple'))

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(table)
        click.echo(f"已导出到 {export}")


def _report_depreciation(conn, export, fmt):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT depreciation_status, category, COUNT(*) as count, SUM(purchase_price) as total_value
        FROM assets
        GROUP BY depreciation_status, category
        ORDER BY depreciation_status, category
    ''')
    rows = cursor.fetchall()

    statuses = {}
    for row in rows:
        status = row['depreciation_status'] or '未设置'
        if status not in statuses:
            statuses[status] = {'总数': 0, '总价值': 0, '类别': {}}
        statuses[status]['总数'] += row['count']
        statuses[status]['总价值'] += row['total_value'] or 0
        statuses[status]['类别'][row['category']] = row['count']

    table = []
    for status, data in statuses.items():
        cat_str = ', '.join(f"{k}:{v}" for k, v in data['类别'].items())
        table.append([status, data['总数'], f"{data['总价值']:.2f}", cat_str])

    headers = ['折旧状态', '数量', '总价值(元)', '类别分布']
    click.echo(tabulate(table, headers=headers, tablefmt='simple'))

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(table)
        click.echo(f"已导出到 {export}")


def _report_value(conn, export, fmt):
    cursor = conn.cursor()

    cursor.execute('SELECT COUNT(*) as cnt, SUM(purchase_price) as val FROM assets')
    total = cursor.fetchone()

    cursor.execute('''
        SELECT category, COUNT(*) as cnt, SUM(purchase_price) as val
        FROM assets GROUP BY category
    ''')
    by_cat = cursor.fetchall()

    cursor.execute('''
        SELECT department, COUNT(*) as cnt, SUM(purchase_price) as val
        FROM assets GROUP BY department
    ''')
    by_dept = cursor.fetchall()

    click.echo("=" * 50)
    click.echo("  资产价值报表")
    click.echo("=" * 50)
    click.echo(f"  总资产: {total['cnt']} 件，总值: {total['val'] or 0:.2f} 元")
    click.echo()
    click.echo("  按类别:")
    for row in by_cat:
        pct = (row['val'] / total['val'] * 100) if total['val'] else 0
        click.echo(f"    {row['category']}: {row['cnt']} 件, {row['val'] or 0:.2f} 元 ({pct:.1f}%)")
    click.echo()
    click.echo("  按部门:")
    for row in by_dept:
        dept = row['department'] or '未分配'
        pct = (row['val'] / total['val'] * 100) if total['val'] else 0
        click.echo(f"    {dept}: {row['cnt']} 件, {row['val'] or 0:.2f} 元 ({pct:.1f}%)")
    click.echo("=" * 50)

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['类别', '数量', '价值(元)', '占比(%)'])
            for row in by_cat:
                pct = (row['val'] / total['val'] * 100) if total['val'] else 0
                writer.writerow([row['category'], row['cnt'], f"{row['val'] or 0:.2f}", f"{pct:.1f}"])
            writer.writerow([])
            writer.writerow(['部门', '数量', '价值(元)', '占比(%)'])
            for row in by_dept:
                dept = row['department'] or '未分配'
                pct = (row['val'] / total['val'] * 100) if total['val'] else 0
                writer.writerow([dept, row['cnt'], f"{row['val'] or 0:.2f}", f"{pct:.1f}"])
        click.echo(f"已导出到 {export}")


@click.command('inventory-diff')
@click.argument('inventory_file', type=click.Path(exists=True))
@click.option('--column', default='asset_no', help='盘点文件中的资产编号列名')
@click.option('--export', '-e', help='导出差异报告')
def inventory_diff_cmd(inventory_file, column, export):
    """生成盘点差异报告"""
    import csv

    inventory_nos = set()
    ext = inventory_file.rsplit('.', 1)[-1].lower()

    if ext == 'csv':
        with open(inventory_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                val = row.get(column) or row.get('资产编号') or row.get('编号')
                if val:
                    inventory_nos.add(val.strip())
    elif ext in ['xlsx', 'xls']:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(inventory_file, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            col_idx = None
            for i, h in enumerate(headers):
                if h and str(h).strip() in [column, '资产编号', '编号']:
                    col_idx = i
                    break
            if col_idx is None:
                col_idx = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if col_idx < len(row) and row[col_idx]:
                    inventory_nos.add(str(row[col_idx]).strip())
        except ImportError:
            click.echo("错误: 需要 openpyxl 库读取 Excel 文件")
            return
    else:
        with open(inventory_file, 'r', encoding='utf-8-sig') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    inventory_nos.add(line)

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT asset_no, name, category, status, department FROM assets')
        db_assets = {row['asset_no']: row for row in cursor.fetchall()}

    db_nos = set(db_assets.keys())

    missing = db_nos - inventory_nos
    extra = inventory_nos - db_nos
    matched = db_nos & inventory_nos

    click.echo("=" * 60)
    click.echo("  盘点差异报告")
    click.echo("=" * 60)
    click.echo(f"  系统资产数: {len(db_nos)}")
    click.echo(f"  盘点资产数: {len(inventory_nos)}")
    click.echo(f"  盘盈(系统没有): {len(extra)}")
    click.echo(f"  盘亏(盘点没有): {len(missing)}")
    click.echo(f"  一致: {len(matched)}")
    click.echo("=" * 60)

    if missing:
        click.echo(f"\n盘亏资产 ({len(missing)} 项):")
        table = []
        for no in sorted(missing):
            asset = db_assets[no]
            table.append([
                asset['asset_no'],
                asset['name'],
                asset['category'],
                asset['status'],
                asset['department'] or '-'
            ])
        click.echo(tabulate(table, headers=['资产编号', '名称', '类别', '状态', '部门'], tablefmt='simple'))

    if extra:
        click.echo(f"\n盘盈资产 ({len(extra)} 项):")
        for no in sorted(extra):
            click.echo(f"  {no}")

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['盘点差异报告'])
            writer.writerow(['系统资产数', len(db_nos)])
            writer.writerow(['盘点资产数', len(inventory_nos)])
            writer.writerow(['盘盈', len(extra)])
            writer.writerow(['盘亏', len(missing)])
            writer.writerow([])

            writer.writerow(['盘亏资产'])
            writer.writerow(['资产编号', '名称', '类别', '状态', '部门'])
            for no in sorted(missing):
                asset = db_assets[no]
                writer.writerow([
                    asset['asset_no'], asset['name'], asset['category'],
                    asset['status'], asset['department'] or ''
                ])
            writer.writerow([])

            writer.writerow(['盘盈资产'])
            for no in sorted(extra):
                writer.writerow([no])
        click.echo(f"\n差异报告已导出到 {export}")


@click.command('history')
@click.option('--asset-no', '-n', help='按资产编号筛选')
@click.option('--operation', '-o', help='按操作类型筛选')
@click.option('--operator', help='按操作人筛选')
@click.option('--date-from', help='开始日期 (YYYY-MM-DD)')
@click.option('--date-to', help='结束日期 (YYYY-MM-DD)')
@click.option('--limit', type=int, default=50, help='显示条数')
@click.option('--export', '-e', help='导出到 CSV 文件')
def history_cmd(asset_no, operation, operator, date_from, date_to, limit, export):
    """查看操作历史"""
    filters = {}
    if asset_no:
        filters['asset_no'] = asset_no
    if operation:
        filters['operation'] = operation
    if date_from:
        filters['date_from'] = date_from
    if date_to:
        filters['date_to'] = date_to

    with get_db() as conn:
        logs = get_operation_logs(conn, filters)

    if operator:
        logs = [l for l in logs if operator in (l['operator'] or '')]

    if not logs:
        click.echo("没有操作记录")
        return

    if limit and limit > 0:
        logs = logs[:limit]

    table = []
    for log in logs:
        table.append([
            log['created_at'],
            log['asset_no'],
            log['operation'],
            log['operator'] or '-',
            log['detail'] or '-'
        ])

    headers = ['时间', '资产编号', '操作', '操作人', '详情']
    click.echo(tabulate(table, headers=headers, tablefmt='simple'))
    click.echo(f"\n共 {len(logs)} 条记录")

    if export:
        with open(export, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(table)
        click.echo(f"已导出到 {export}")


@click.command('export')
@click.argument('output_file')
@click.option('--category', '-c', help='按类别导出')
@click.option('--department', '-d', help='按部门导出')
@click.option('--status', '-s', type=click.Choice(ASSET_STATUS), help='按状态导出')
@click.option('--date-from', help='购入日期起始')
@click.option('--date-to', help='购入日期截止')
@click.option('--format', '-f', 'fmt', type=click.Choice(['csv', 'xlsx']), default='csv',
              help='导出格式')
def export_cmd(output_file, category, department, status, date_from, date_to, fmt):
    """按条件导出资产报表"""
    filters = {}
    if category:
        filters['category'] = category
    if department:
        filters['department'] = department
    if status:
        filters['status'] = status
    if date_from:
        filters['date_from'] = date_from
    if date_to:
        filters['date_to'] = date_to

    with get_db() as conn:
        assets = query_assets(conn, filters)

    if not assets:
        click.echo("没有符合条件的资产")
        return

    headers = ['资产编号', '资产名称', '类别', '品牌', '型号', '序列号',
               '购入日期', '购入价格', '所属部门', '存放地点', '使用人',
               '状态', '折旧状态', '备注', '创建时间', '更新时间']
    keys = ['asset_no', 'name', 'category', 'brand', 'model', 'serial_no',
            'purchase_date', 'purchase_price', 'department', 'location', 'user_name',
            'status', 'depreciation_status', 'remark', 'created_at', 'updated_at']

    rows = []
    for asset in assets:
        row = []
        for k in keys:
            val = asset[k]
            if k == 'purchase_price' and val:
                val = f"{val:.2f}"
            row.append(val if val is not None else '')
        rows.append(row)

    if fmt == 'csv':
        with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
    elif fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = '资产清单'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            wb.save(output_file)
        except ImportError:
            click.echo("错误: 需要安装 openpyxl 才能导出 Excel 文件")
            return

    click.echo(f"已导出 {len(assets)} 条资产到 {output_file}")


@click.command('monthly-report')
@click.option('--date-from', help='开始月份 (YYYY-MM)')
@click.option('--date-to', help='结束月份 (YYYY-MM)')
@click.option('--months', type=int, help='最近 N 个月')
@click.option('--export', '-e', help='导出到文件')
@click.option('--format', '-f', 'fmt', type=click.Choice(['csv', 'xlsx']), default='csv',
              help='导出格式')
@click.option('--detail', is_flag=True, help='显示详细分类数据')
def monthly_report_cmd(date_from, date_to, months, export, fmt, detail):
    """月度资产变化统计报表"""
    from datetime import datetime, timedelta

    now = datetime.now()

    if months and not date_from:
        start_date = now - timedelta(days=months * 30)
        date_from = start_date.strftime('%Y-%m')
    elif not date_from:
        date_from = '2020-01'

    if not date_to:
        date_to = now.strftime('%Y-%m')

    start = datetime.strptime(date_from, '%Y-%m')
    end = datetime.strptime(date_to, '%Y-%m')

    all_months = []
    m = start
    while m <= end:
        all_months.append(m.strftime('%Y-%m'))
        if m.month == 12:
            m = datetime(m.year + 1, 1, 1)
        else:
            m = datetime(m.year, m.month + 1, 1)

    with get_db() as conn:
        cursor = conn.cursor()

        cursor.execute('''
            SELECT strftime('%Y-%m', created_at) as month,
                   COUNT(*) as count,
                   SUM(purchase_price) as value
            FROM assets
            WHERE strftime('%Y-%m', created_at) >= ? AND strftime('%Y-%m', created_at) <= ?
            GROUP BY month
            ORDER BY month
        ''', (date_from, date_to))
        new_assets = {row['month']: {'count': row['count'], 'value': row['value'] or 0}
                      for row in cursor.fetchall()}

        operation_stats = {}
        op_types = ['分配', '归还', '维修发起', '维修完成', '报废', '闲置', '导入更新']
        for op in op_types:
            cursor.execute('''
                SELECT strftime('%Y-%m', created_at) as month,
                       COUNT(*) as count
                FROM operation_logs
                WHERE operation = ?
                  AND strftime('%Y-%m', created_at) >= ?
                  AND strftime('%Y-%m', created_at) <= ?
                GROUP BY month
                ORDER BY month
            ''', (op, date_from, date_to))
            operation_stats[op] = {row['month']: row['count'] for row in cursor.fetchall()}

        cursor.execute('''
            SELECT strftime('%Y-%m', ol.created_at) as month,
                   ol.detail
            FROM operation_logs ol
            WHERE ol.operation = '维修完成'
              AND strftime('%Y-%m', ol.created_at) >= ?
              AND strftime('%Y-%m', ol.created_at) <= ?
        ''', (date_from, date_to))
        repair_cost_rows = cursor.fetchall()

        repair_cost_by_month = {}
        import re
        cost_pattern = re.compile(r'费用[:：]\s*([\d.]+)')
        for row in repair_cost_rows:
            cost = 0
            match = cost_pattern.search(row['detail'] or '')
            if match:
                cost = float(match.group(1))
            month = row['month']
            repair_cost_by_month[month] = repair_cost_by_month.get(month, 0) + cost

        cursor.execute('''
            SELECT strftime('%Y-%m', ol.created_at) as month,
                   a.purchase_price
            FROM operation_logs ol
            JOIN assets a ON ol.asset_no = a.asset_no
            WHERE ol.operation = '报废'
              AND strftime('%Y-%m', ol.created_at) >= ?
              AND strftime('%Y-%m', ol.created_at) <= ?
        ''', (date_from, date_to))
        scrap_value_rows = cursor.fetchall()

        scrap_value_by_month = {}
        for row in scrap_value_rows:
            month = row['month']
            scrap_value_by_month[month] = scrap_value_by_month.get(month, 0) + (row['purchase_price'] or 0)

        cursor.execute('SELECT COUNT(*) as cnt, SUM(purchase_price) as val FROM assets')
        total_row = cursor.fetchone()
        current_total = total_row['cnt']
        current_value = total_row['val'] or 0

        start_month = all_months[0]
        cursor.execute('''
            SELECT COUNT(*) as cnt, SUM(purchase_price) as val
            FROM assets
            WHERE strftime('%Y-%m', created_at) < ?
        ''', (start_month,))
        before_row = cursor.fetchone()
        start_count = before_row['cnt'] or 0
        start_value = before_row['val'] or 0

    click.echo("=" * 100)
    click.echo("  月度资产变化统计报表")
    click.echo(f"  统计区间: {date_from} ~ {date_to}")
    click.echo("=" * 100)

    headers = ['月份', '新增数', '新增价值', '报废数', '报废价值',
               '净增数', '净价值变化', '期末数', '期末总值',
               '分配', '归还', '维修费用']

    rows = []
    running_count = start_count
    running_value = start_value

    for month in all_months:
        new_count = new_assets.get(month, {}).get('count', 0)
        new_val = new_assets.get(month, {}).get('value', 0)
        scrap_count = operation_stats.get('报废', {}).get(month, 0)
        scrap_val = scrap_value_by_month.get(month, 0)
        assign_count = operation_stats.get('分配', {}).get(month, 0)
        return_count = operation_stats.get('归还', {}).get(month, 0)
        repair_cost = repair_cost_by_month.get(month, 0)

        net_count = new_count - scrap_count
        net_value = new_val - scrap_val

        running_count += net_count
        running_value += net_value

        rows.append([
            month,
            new_count,
            f"{new_val:,.2f}",
            scrap_count,
            f"{scrap_val:,.2f}",
            f"{net_count:+d}",
            f"{net_value:+,.2f}",
            running_count,
            f"{running_value:,.2f}",
            assign_count,
            return_count,
            f"{repair_cost:,.2f}",
        ])

    click.echo(tabulate(rows, headers=headers, tablefmt='simple'))

    click.echo()
    click.echo("-" * 100)
    total_new = sum(new_assets.get(m, {}).get('count', 0) for m in all_months)
    total_new_val = sum(new_assets.get(m, {}).get('value', 0) for m in all_months)
    total_assign = sum(operation_stats.get('分配', {}).get(m, 0) for m in all_months)
    total_return = sum(operation_stats.get('归还', {}).get(m, 0) for m in all_months)
    total_scrap = sum(operation_stats.get('报废', {}).get(m, 0) for m in all_months)
    total_scrap_val = sum(scrap_value_by_month.get(m, 0) for m in all_months)
    total_repair_cost = sum(repair_cost_by_month.get(m, 0) for m in all_months)
    total_net_count = total_new - total_scrap
    total_net_value = total_new_val - total_scrap_val

    click.echo(f"  期初资产: {start_count} 件 / {start_value:,.2f} 元")
    click.echo(f"  本期新增: {total_new} 件 / {total_new_val:,.2f} 元")
    click.echo(f"  本期报废: {total_scrap} 件 / {total_scrap_val:,.2f} 元")
    click.echo(f"  本期净增: {total_net_count:+d} 件 / {total_net_value:+,.2f} 元")
    click.echo(f"  期末资产: {current_total} 件 / {current_value:,.2f} 元")
    click.echo(f"  本期分配: {total_assign} 次，归还: {total_return} 次")
    click.echo(f"  本期维修费用: {total_repair_cost:,.2f} 元")
    click.echo("=" * 100)

    if export:
        if fmt == 'csv':
            with open(export, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['月度资产变化统计报表', f'{date_from} ~ {date_to}'])
                writer.writerow([])
                writer.writerow(headers)
                writer.writerows(rows)
                writer.writerow([])
                writer.writerow(['汇总统计'])
                writer.writerow(['期初资产数', start_count])
                writer.writerow(['期初资产总值(元)', f"{start_value:.2f}"])
                writer.writerow(['新增数量', total_new])
                writer.writerow(['新增价值(元)', f"{total_new_val:.2f}"])
                writer.writerow(['报废数量', total_scrap])
                writer.writerow(['报废价值(元)', f"{total_scrap_val:.2f}"])
                writer.writerow(['净增数量', total_net_count])
                writer.writerow(['净价值变化(元)', f"{total_net_value:.2f}"])
                writer.writerow(['期末资产数', current_total])
                writer.writerow(['期末资产总值(元)', f"{current_value:.2f}"])
                writer.writerow(['分配次数', total_assign])
                writer.writerow(['归还次数', total_return])
                writer.writerow(['维修费用(元)', f"{total_repair_cost:.2f}"])
        elif fmt == 'xlsx':
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = '月度变化'
                ws.append(['月度资产变化统计报表', f'{date_from} ~ {date_to}'])
                ws.append([])
                ws.append(headers)
                for row in rows:
                    ws.append(row)
                ws.append([])
                ws.append(['汇总统计'])
                ws.append(['期初资产数', start_count])
                ws.append(['期初资产总值(元)', start_value])
                ws.append(['新增数量', total_new])
                ws.append(['新增价值(元)', total_new_val])
                ws.append(['报废数量', total_scrap])
                ws.append(['报废价值(元)', total_scrap_val])
                ws.append(['净增数量', total_net_count])
                ws.append(['净价值变化(元)', total_net_value])
                ws.append(['期末资产数', current_total])
                ws.append(['期末资产总值(元)', current_value])
                ws.append(['分配次数', total_assign])
                ws.append(['归还次数', total_return])
                ws.append(['维修费用(元)', total_repair_cost])
                wb.save(export)
            except ImportError:
                click.echo("错误: 需要安装 openpyxl 才能导出 Excel 文件")
                return

        click.echo(f"\n报表已导出到: {export}")
