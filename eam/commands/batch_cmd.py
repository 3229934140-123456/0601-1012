import os
import csv
import click
from datetime import datetime
from tabulate import tabulate
from ..database import (
    get_db, get_asset_by_no, update_asset, log_operation,
    update_asset_timestamp, create_batch_run, update_batch_run,
    save_snapshot, get_batch_run, get_batch_snapshots, list_batch_runs
)


SUPPORTED_OPERATIONS = ['assign', 'return', 'move', 'scrap', 'remark', 'idle', 'status']


def _read_tasks(filepath):
    """读取批量任务文件"""
    ext = os.path.splitext(filepath)[1].lower()
    tasks = []

    if ext == '.csv':
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader, 2):
                task = {}
                for k, v in row.items():
                    if k is not None and v is not None:
                        task[k.strip()] = v.strip()
                task['_line'] = i
                tasks.append(task)
    elif ext in ['.xlsx', '.xls']:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                task = {}
                for j, val in enumerate(row):
                    if j < len(headers) and headers[j]:
                        task[headers[j]] = str(val).strip() if val is not None else ''
                task['_line'] = i
                if task.get('operation') or task.get('操作'):
                    tasks.append(task)
        except ImportError:
            click.echo("错误: 需要安装 openpyxl 才能读取 Excel 文件")
            return []
    else:
        click.echo(f"错误: 不支持的文件格式 {ext}")
        return []

    return tasks


def _normalize_task(raw):
    """标准化任务字段"""
    mapping = {
        '操作': 'operation',
        '操作类型': 'operation',
        '建议批量操作': 'operation',
        '资产编号': 'asset_no',
        '编号': 'asset_no',
        '使用人': 'user_name',
        '保管人': 'user_name',
        '部门': 'department',
        '所属部门': 'department',
        '地点': 'location',
        '存放地点': 'location',
        '新地点': 'location',
        '状态': 'status',
        '新状态': 'status',
        '备注': 'remark',
        '建议备注': 'remark',
        '原因': 'reason',
        '报废原因': 'reason',
    }

    task = {}
    for k, v in raw.items():
        key = mapping.get(k, k)
        task[key] = v

    return task


def _is_audit_csv(tasks):
    """判断是否是审计导出的待处理清单 CSV"""
    if not tasks:
        return False
    first = tasks[0]
    audit_keys = ['建议批量操作', '建议备注', '异常类型', '问题描述']
    found = sum(1 for k in audit_keys if k in first)
    return found >= 2


def _row_to_dict(row):
    """将 sqlite3.Row 转为 dict"""
    if row is None:
        return None
    return {k: row[k] for k in row.keys()}


def _validate_task(task, conn):
    """验证单个任务，返回 (is_valid, error_message, asset_dict)"""
    op = task.get('operation', '').lower()
    asset_no = task.get('asset_no', '')

    if not op:
        return False, '缺少操作类型', None

    if op not in SUPPORTED_OPERATIONS:
        return False, f'不支持的操作类型: {op}', None

    if not asset_no:
        return False, '缺少资产编号', None

    asset = get_asset_by_no(conn, asset_no)
    if not asset:
        return False, f'资产不存在: {asset_no}', None

    asset_dict = _row_to_dict(asset)

    if op == 'assign':
        if not task.get('user_name'):
            return False, '分配操作缺少使用人', asset_dict

    if op == 'move':
        if not task.get('location'):
            return False, '移动操作缺少新地点', asset_dict

    if op == 'status':
        if not task.get('status'):
            return False, '状态变更操作缺少新状态', asset_dict
        valid_statuses = ['闲置', '在用', '维修中', '已报废']
        if task['status'] not in valid_statuses:
            return False, f'无效状态: {task["status"]}', asset_dict

    if op == 'remark':
        if not task.get('remark'):
            return False, '备注操作缺少备注内容', asset_dict

    return True, '', asset_dict


def _execute_task(task, asset, conn, operator):
    """执行单个任务，返回 (success, message)"""
    op = task['operation'].lower()
    asset_no = task['asset_no']
    remark = task.get('remark', '') or ''
    reason = task.get('reason', '') or ''

    try:
        if op == 'assign':
            user_name = task['user_name']
            department = task.get('department', '') or asset.get('department') or ''
            location = task.get('location', '') or asset.get('location') or ''

            update_data = {
                'user_name': user_name,
                'status': '在用',
            }
            if department:
                update_data['department'] = department
            if location:
                update_data['location'] = location

            if remark:
                old_remark = asset['remark'] or ''
                update_data['remark'] = old_remark + ('; ' if old_remark else '') + remark

            update_asset(conn, asset_no, update_data)
            update_asset_timestamp(conn, asset_no)

            detail = f"分配给 {user_name}"
            if department:
                detail += f"（{department}）"
            if remark:
                detail += f" | 备注: {remark}"

            log_operation(conn, asset_no, '分配', operator, detail)
            return True, '分配成功'

        elif op == 'return':
            old_user = asset['user_name'] or '未知'

            update_data = {'user_name': '', 'status': '闲置'}
            if remark:
                old_remark = asset['remark'] or ''
                update_data['remark'] = old_remark + ('; ' if old_remark else '') + remark

            update_asset(conn, asset_no, update_data)
            update_asset_timestamp(conn, asset_no)

            detail = f"归还，原使用人: {old_user}"
            if remark:
                detail += f" | 备注: {remark}"

            log_operation(conn, asset_no, '归还', operator, detail)
            return True, '归还成功'

        elif op == 'move':
            new_location = task['location']
            old_location = asset['location'] or '未设置'

            update_data = {'location': new_location}
            if remark:
                old_remark = asset['remark'] or ''
                update_data['remark'] = old_remark + ('; ' if old_remark else '') + remark

            update_asset(conn, asset_no, update_data)
            update_asset_timestamp(conn, asset_no)

            detail = f"从 {old_location} 移动到 {new_location}"
            if remark:
                detail += f" | 备注: {remark}"

            log_operation(conn, asset_no, '移动', operator, detail)
            return True, '移动成功'

        elif op == 'scrap':
            update_data = {'status': '已报废', 'user_name': ''}
            scrap_reason = reason or remark or ''
            if scrap_reason:
                old_remark = asset['remark'] or ''
                update_data['remark'] = old_remark + ('; ' if old_remark else '') + f'报废原因: {scrap_reason}'

            update_asset(conn, asset_no, update_data)
            update_asset_timestamp(conn, asset_no)

            detail = '资产报废'
            if scrap_reason:
                detail += f' | 原因: {scrap_reason}'

            log_operation(conn, asset_no, '报废', operator, detail)
            return True, '报废成功'

        elif op == 'remark':
            if not remark:
                return False, '备注操作缺少备注内容'

            old_remark = asset['remark'] or ''
            new_remark = old_remark + ('; ' if old_remark else '') + remark

            update_asset(conn, asset_no, {'remark': new_remark})
            update_asset_timestamp(conn, asset_no)

            log_operation(conn, asset_no, '备注更新', operator, f'追加备注: {remark}')
            return True, '备注追加成功'

        elif op == 'idle':
            old_status = asset['status']
            old_user = asset['user_name'] or '无'

            update_data = {'status': '闲置', 'user_name': ''}
            if remark:
                old_remark = asset['remark'] or ''
                update_data['remark'] = old_remark + ('; ' if old_remark else '') + f'闲置原因: {remark}'

            update_asset(conn, asset_no, update_data)
            update_asset_timestamp(conn, asset_no)

            detail = f"从 {old_status} 标记为闲置，原使用人: {old_user}"
            if remark:
                detail += f" | 原因: {remark}"

            log_operation(conn, asset_no, '闲置', operator, detail)
            return True, '标记闲置成功'

        elif op == 'status':
            new_status = task['status']
            old_status = asset['status']

            update_data = {'status': new_status}
            if new_status in ['闲置', '已报废']:
                update_data['user_name'] = ''
            if remark:
                old_remark = asset['remark'] or ''
                update_data['remark'] = old_remark + ('; ' if old_remark else '') + remark

            update_asset(conn, asset_no, update_data)
            update_asset_timestamp(conn, asset_no)

            detail = f"状态从 {old_status} 变更为 {new_status}"
            if remark:
                detail += f" | 备注: {remark}"

            log_operation(conn, asset_no, '状态变更', operator, detail)
            return True, '状态变更成功'

        else:
            return False, f'未知操作: {op}'

    except Exception as e:
        return False, f'执行异常: {str(e)}'


@click.command('batch')
@click.argument('filepath', type=click.Path(exists=True))
@click.option('--dry-run', is_flag=True, help='预览模式，仅校验不执行')
@click.option('--operator', default='system', help='操作人')
@click.option('--yes', '-y', is_flag=True, help='跳过确认')
@click.option('--fail-log', help='失败记录导出文件')
@click.option('--remark', help='批次备注')
def batch_cmd(filepath, dry_run, operator, yes, fail_log, remark):
    """批量执行任务（支持分配、归还、移动、报废、备注等）

    \b
    CSV 文件示例:
    operation,asset_no,user_name,department,location,remark
    assign,PC-001,张三,研发部,,
    return,PC-002,,,,
    move,PC-003,,,3楼会议室,
    scrap,PC-004,,,老旧损坏
    remark,PC-005,,,,新增备注内容
    idle,BADGE-001,,,离职人员
    status,OF-001,,,闲置,
    """
    raw_tasks = _read_tasks(filepath)
    if not raw_tasks:
        click.echo("未读取到任何任务")
        return

    is_audit_source = _is_audit_csv(raw_tasks)
    if is_audit_source:
        click.echo("检测到审计待处理清单，将按建议批量操作执行\n")

    click.echo(f"读取到 {len(raw_tasks)} 条任务，正在校验...\n")

    tasks = []
    for raw in raw_tasks:
        task = _normalize_task(raw)
        task['_line'] = raw.get('_line', '?')
        task['_raw'] = raw
        tasks.append(task)

    with get_db() as conn:
        valid_tasks = []
        failed_tasks = []

        for task in tasks:
            is_valid, err_msg, asset = _validate_task(task, conn)
            task['_asset'] = asset
            task['_valid'] = is_valid
            task['_error'] = err_msg

            if is_valid:
                valid_tasks.append(task)
            else:
                failed_tasks.append(task)

    op_counts = {}
    for task in valid_tasks:
        op = task['operation'].lower()
        op_counts[op] = op_counts.get(op, 0) + 1

    click.echo("=" * 60)
    click.echo("  任务校验结果")
    click.echo("=" * 60)
    click.echo(f"  总任务数: {len(tasks)}")
    click.echo(f"  校验通过: {len(valid_tasks)}")
    click.echo(f"  校验失败: {len(failed_tasks)}")
    click.echo()
    click.echo("  操作类型分布:")
    for op, cnt in sorted(op_counts.items()):
        click.echo(f"    {op}: {cnt} 条")

    if failed_tasks:
        click.echo("\n" + "-" * 60)
        click.echo("  失败明细:")
        click.echo("-" * 60)
        rows = []
        for t in failed_tasks[:20]:
            rows.append([
                t.get('_line', '?'),
                t.get('asset_no', ''),
                t.get('operation', ''),
                t.get('_error', '')
            ])
        click.echo(tabulate(rows, headers=['行号', '资产编号', '操作', '失败原因'],
                            tablefmt='simple'))
        if len(failed_tasks) > 20:
            click.echo(f"  ... 还有 {len(failed_tasks) - 20} 条失败记录")

    if dry_run:
        click.echo("\n" + "=" * 60)
        click.echo("  [预览模式] 以下任务将会执行:")
        click.echo("=" * 60)

        if valid_tasks:
            rows = []
            for t in valid_tasks[:20]:
                asset = t['_asset']
                rows.append([
                    t.get('_line', '?'),
                    t['asset_no'],
                    t['operation'],
                    asset['name'] if asset else '',
                    _task_summary(t, asset)
                ])
            click.echo(tabulate(rows, headers=['行号', '资产编号', '操作', '资产名称', '操作摘要'],
                                tablefmt='simple'))
            if len(valid_tasks) > 20:
                click.echo(f"  ... 还有 {len(valid_tasks) - 20} 条")

        click.echo("\n[预览模式] 未实际执行任何操作")
        return

    if not valid_tasks:
        click.echo("\n没有可执行的任务")
        return

    if not yes:
        click.echo(f"\n将执行 {len(valid_tasks)} 条任务")
        if not click.confirm("确认执行?", default=True):
            click.echo("已取消")
            return

    batch_no = 'B' + datetime.now().strftime('%Y%m%d%H%M%S')
    success_count = 0
    exec_failed = []

    batch_remark = remark or ''
    if is_audit_source:
        audit_prefix = '[审计处理]'
        if batch_remark:
            batch_remark = audit_prefix + ' ' + batch_remark
        else:
            batch_remark = audit_prefix

    with get_db() as conn:
        create_batch_run(conn, batch_no, operator, len(valid_tasks), batch_remark)

        snapshotted = set()
        for task in valid_tasks:
            asset_no = task['asset_no']
            if asset_no in snapshotted:
                continue
            asset_row = get_asset_by_no(conn, asset_no)
            if asset_row:
                asset_dict = _row_to_dict(asset_row)
                save_snapshot(conn, batch_no, asset_no, asset_dict)
                snapshotted.add(asset_no)

        for task in valid_tasks:
            asset_row = get_asset_by_no(conn, task['asset_no'])
            if not asset_row:
                task['_exec_error'] = '资产不存在'
                exec_failed.append(task)
                continue
            asset = _row_to_dict(asset_row)

            success, msg = _execute_task(task, asset, conn, operator)
            if success:
                success_count += 1
            else:
                task['_exec_error'] = msg
                exec_failed.append(task)

        fail_count = len(exec_failed) + len(failed_tasks)
        update_batch_run(conn, batch_no, success_count, fail_count, 'completed')

    total_failed = len(failed_tasks) + len(exec_failed)

    click.echo("\n" + "=" * 60)
    click.echo("  执行结果")
    click.echo("=" * 60)
    click.echo(f"  批次号: {batch_no}")
    click.echo(f"  成功: {success_count} 条")
    click.echo(f"  失败: {total_failed} 条（校验失败 {len(failed_tasks)} + 执行失败 {len(exec_failed)}）")
    click.echo()
    click.echo(f"  提示: 使用 batch-rollback {batch_no} 可撤回本次操作")

    if exec_failed:
        click.echo("\n执行失败明细:")
        rows = []
        for t in exec_failed:
            rows.append([
                t.get('_line', '?'),
                t['asset_no'],
                t['operation'],
                t.get('_exec_error', '')
            ])
        click.echo(tabulate(rows, headers=['行号', '资产编号', '操作', '失败原因'],
                            tablefmt='simple'))

    if fail_log and (failed_tasks or exec_failed):
        with open(fail_log, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['行号', '资产编号', '操作', '失败类型', '失败原因'])
            for t in failed_tasks:
                writer.writerow([
                    t.get('_line', '?'),
                    t.get('asset_no', ''),
                    t.get('operation', ''),
                    '校验失败',
                    t.get('_error', '')
                ])
            for t in exec_failed:
                writer.writerow([
                    t.get('_line', '?'),
                    t.get('asset_no', ''),
                    t.get('operation', ''),
                    '执行失败',
                    t.get('_exec_error', '')
                ])
        click.echo(f"\n失败记录已导出到: {fail_log}")


def _task_summary(task, asset):
    """生成任务摘要文本"""
    op = task['operation'].lower()
    if op == 'assign':
        return f"分配给 {task.get('user_name', '')}"
    elif op == 'return':
        old_user = asset['user_name'] or '无' if asset else '?'
        return f"归还（使用人: {old_user}）"
    elif op == 'move':
        old_loc = asset['location'] or '无' if asset else '?'
        return f"{old_loc} -> {task.get('location', '')}"
    elif op == 'scrap':
        return '报废'
    elif op == 'remark':
        return f"追加备注: {task.get('remark', '')[:20]}"
    elif op == 'idle':
        old_status = asset['status'] if asset else '?'
        return f"{old_status} -> 闲置"
    elif op == 'status':
        old_status = asset['status'] if asset else '?'
        return f"{old_status} -> {task.get('status', '')}"
    return ''


@click.command('batch-template')
@click.option('--output', '-o', default='batch_template.csv', help='输出文件名')
def batch_template_cmd(output):
    """生成批量任务 CSV 模板"""
    headers = [
        'operation', 'asset_no', 'user_name', 'department',
        'location', 'status', 'remark', 'reason'
    ]
    examples = [
        ['assign', 'PC-001', '张三', '研发部', '', '', '', ''],
        ['return', 'PC-002', '', '', '', '', '', ''],
        ['move', 'PC-003', '', '', '3楼会议室', '', '', ''],
        ['scrap', 'PC-004', '', '', '', '', '', '老旧损坏'],
        ['remark', 'PC-005', '', '', '', '', '新增备注内容', ''],
        ['idle', 'BADGE-001', '', '', '', '', '离职人员', ''],
        ['status', 'OF-001', '', '', '', '闲置', '', ''],
    ]

    import csv
    with open(output, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(examples)

    click.echo(f"模板已生成: {output}")
    click.echo("\n支持的操作类型:")
    click.echo("  assign   - 分配资产 (需 user_name)")
    click.echo("  return   - 归还资产")
    click.echo("  move     - 移动资产 (需 location)")
    click.echo("  scrap    - 报废资产")
    click.echo("  remark   - 追加备注 (需 remark)")
    click.echo("  idle     - 标记闲置")
    click.echo("  status   - 变更状态 (需 status)")


@click.command('batch-list')
@click.option('--limit', '-n', type=int, default=10, help='显示最近 N 条')
def batch_list_cmd(limit):
    """查看批量任务历史"""
    with get_db() as conn:
        runs = list_batch_runs(conn, limit)

    if not runs:
        click.echo("暂无批量任务记录")
        return

    rows = []
    for r in runs:
        rows.append([
            r['batch_no'],
            r['status'],
            r['operator'] or '-',
            r['task_count'],
            r['success_count'] or 0,
            r['fail_count'] or 0,
            r['created_at'],
            r['remark'] or '',
        ])

    click.echo(tabulate(rows,
                        headers=['批次号', '状态', '操作人', '任务数', '成功', '失败', '创建时间', '备注'],
                        tablefmt='simple'))


@click.command('batch-rollback')
@click.argument('batch_no')
@click.option('--operator', default='system', help='操作人')
@click.option('--yes', '-y', is_flag=True, help='跳过确认')
@click.option('--dry-run', is_flag=True, help='预览撤回内容')
def batch_rollback_cmd(batch_no, operator, yes, dry_run):
    """撤回批量任务（按批次号恢复资产状态）"""
    with get_db() as conn:
        batch = get_batch_run(conn, batch_no)
        if not batch:
            click.echo(f"错误: 批次 {batch_no} 不存在")
            return

        if batch['status'] == 'rolled_back':
            click.echo(f"警告: 批次 {batch_no} 已经撤回过")

        snapshots = get_batch_snapshots(conn, batch_no)
        if not snapshots:
            click.echo("错误: 该批次没有快照数据，无法撤回")
            return

    if dry_run:
        click.echo(f"批次 {batch_no} 将撤回 {len(snapshots)} 项资产的变更:")
        click.echo()
        rows = []
        for s in snapshots:
            before = s['before_data']
            rows.append([
                s['asset_no'],
                before.get('status', ''),
                before.get('user_name', '') or '',
                before.get('department', '') or '',
                before.get('location', '') or '',
            ])
        click.echo(tabulate(rows, headers=['资产编号', '原状态', '原使用人', '原部门', '原地点'],
                            tablefmt='simple'))
        click.echo(f"\n[预览] 将恢复以上 {len(snapshots)} 项资产到变更前状态")
        return

    if not yes:
        click.echo(f"将撤回批次 {batch_no} 的 {len(snapshots)} 项资产变更")
        if not click.confirm("确认撤回?", default=False):
            click.echo("已取消")
            return

    with get_db() as conn:
        rollback_count = 0
        for s in snapshots:
            asset_no = s['asset_no']
            before = s['before_data']

            existing = get_asset_by_no(conn, asset_no)
            if not existing:
                continue

            update_data = {}
            for key in ['status', 'user_name', 'department', 'location', 'remark', 'depreciation_status']:
                old_val = before.get(key, '') or ''
                cur_val = existing[key] if key in existing.keys() else ''
                if old_val != (cur_val or ''):
                    update_data[key] = old_val

            if update_data:
                update_asset(conn, asset_no, update_data)
                update_asset_timestamp(conn, asset_no)
                log_operation(
                    conn,
                    asset_no,
                    '撤回',
                    operator,
                    f"批次 {batch_no} 撤回，恢复字段: {', '.join(update_data.keys())}"
                )
                rollback_count += 1

        cursor = conn.cursor()
        cursor.execute('''
            UPDATE batch_runs SET status = 'rolled_back',
                   completed_at = datetime('now', 'localtime')
            WHERE batch_no = ?
        ''', (batch_no,))

    click.echo()
    click.echo("=" * 60)
    click.echo("  撤回完成")
    click.echo("=" * 60)
    click.echo(f"  批次号: {batch_no}")
    click.echo(f"  撤回资产: {rollback_count} 项")
    click.echo(f"  操作人: {operator}")
    click.echo("=" * 60)
