import os
import csv
from collections import Counter
import click
from tabulate import tabulate
from ..database import (
    get_db, asset_exists, insert_asset, update_asset, get_asset_by_no,
    log_operation, ASSET_CATEGORIES, ASSET_STATUS, update_asset_timestamp
)


def _read_csv(filepath):
    assets = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            asset = {k.strip(): v.strip() for k, v in row.items() if v is not None}
            assets.append(asset)
    return assets


def _read_excel(filepath):
    try:
        from openpyxl import load_workbook
    except ImportError:
        click.echo("错误: 需要安装 openpyxl 才能读取 Excel 文件。运行: pip install openpyxl")
        return []

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    assets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        asset = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                key = str(headers[i]).strip()
                asset[key] = str(value).strip() if value is not None else ''
        if asset.get('asset_no') or asset.get('资产编号'):
            assets.append(asset)

    return assets


def _normalize_asset(raw):
    mapping = {
        '资产编号': 'asset_no',
        '资产名称': 'name',
        '名称': 'name',
        '类别': 'category',
        '资产类别': 'category',
        '品牌': 'brand',
        '型号': 'model',
        '序列号': 'serial_no',
        '序列号/ SN': 'serial_no',
        '购入日期': 'purchase_date',
        '购买日期': 'purchase_date',
        '购入价格': 'purchase_price',
        '购买价格': 'purchase_price',
        '价格': 'purchase_price',
        '所属部门': 'department',
        '部门': 'department',
        '存放地点': 'location',
        '地点': 'location',
        '使用人': 'user_name',
        '保管人': 'user_name',
        '状态': 'status',
        '折旧状态': 'depreciation_status',
        '备注': 'remark',
    }

    normalized = {}
    for k, v in raw.items():
        key = mapping.get(k, k)
        if key in ['asset_no', 'name', 'category', 'brand', 'model', 'serial_no',
                   'purchase_date', 'purchase_price', 'department', 'location',
                   'user_name', 'status', 'depreciation_status', 'remark']:
            normalized[key] = v

    if not normalized.get('asset_no'):
        return None

    if not normalized.get('name'):
        normalized['name'] = normalized['asset_no']

    if not normalized.get('category'):
        normalized['category'] = '办公设备'

    if not normalized.get('status'):
        normalized['status'] = '闲置'

    if normalized.get('purchase_price'):
        try:
            normalized['purchase_price'] = float(normalized['purchase_price'])
        except (ValueError, TypeError):
            normalized['purchase_price'] = 0
    else:
        normalized['purchase_price'] = 0

    return normalized


def _find_file_duplicates(normalized_list):
    """检查文件内部的重复编号"""
    nos = [a['asset_no'] for a in normalized_list if a.get('asset_no')]
    counter = Counter(nos)
    duplicates = {no: count for no, count in counter.items() if count > 1}
    return duplicates


@click.command('import')
@click.argument('filepath', type=click.Path(exists=True))
@click.option('--dry-run', is_flag=True, help='仅校验不导入')
@click.option('--operator', default='system', help='操作人')
@click.option('--mode', type=click.Choice(['skip', 'merge-remark', 'abort']),
              default='skip',
              help='重复处理模式: skip=跳过, merge-remark=合并备注更新, abort=中止导入')
@click.option('--check-file-dup/--no-check-file-dup', default=True,
              help='是否检查文件内部重复')
def import_cmd(filepath, dry_run, operator, mode, check_file_dup):
    """导入资产清单（支持 CSV 和 Excel）"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.csv':
        raw_assets = _read_csv(filepath)
    elif ext in ['.xlsx', '.xls']:
        raw_assets = _read_excel(filepath)
    else:
        click.echo(f"错误: 不支持的文件格式 {ext}")
        return

    if not raw_assets:
        click.echo("未读取到任何资产数据")
        return

    click.echo(f"读取到 {len(raw_assets)} 条数据，正在处理...")

    normalized_list = []
    invalid_list = []
    for i, raw in enumerate(raw_assets, 1):
        normalized = _normalize_asset(raw)
        if not normalized:
            invalid_list.append((i, '缺少资产编号', raw))
        else:
            normalized_list.append(normalized)

    if invalid_list:
        click.echo(f"\n警告: {len(invalid_list)} 条数据无效:")
        for idx, reason, raw in invalid_list[:5]:
            click.echo(f"  第 {idx} 行: {reason} - {raw}")
        if len(invalid_list) > 5:
            click.echo(f"  ... 还有 {len(invalid_list) - 5} 条")

    file_duplicates = {}
    if check_file_dup:
        file_duplicates = _find_file_duplicates(normalized_list)
        if file_duplicates:
            click.echo(f"\n发现 {len(file_duplicates)} 个文件内部重复编号:")
            table = [[no, cnt] for no, cnt in sorted(file_duplicates.items())]
            click.echo(tabulate(table, headers=['资产编号', '出现次数'], tablefmt='simple'))

            if mode == 'abort':
                click.echo("\n中止导入（文件内存在重复编号）")
                return
            else:
                click.echo(f"\n提示: 文件内重复的编号将只导入第一次出现的记录")

    with get_db() as conn:
        db_duplicates = []
        new_assets = []
        update_assets = []

        seen_nos = set()
        deduped_list = []
        for asset in normalized_list:
            if asset['asset_no'] in seen_nos:
                continue
            seen_nos.add(asset['asset_no'])
            deduped_list.append(asset)

        for asset in deduped_list:
            if asset_exists(conn, asset['asset_no']):
                db_duplicates.append(asset['asset_no'])
                if mode == 'merge-remark':
                    update_assets.append(asset)
            else:
                new_assets.append(asset)

        if db_duplicates:
            click.echo(f"\n发现 {len(db_duplicates)} 个与数据库重复的编号:")
            for no in db_duplicates[:10]:
                click.echo(f"  {no}")
            if len(db_duplicates) > 10:
                click.echo(f"  ... 还有 {len(db_duplicates) - 10} 个")

            if mode == 'abort':
                click.echo("\n中止导入（数据库中存在重复编号）")
                return

        click.echo(f"\n导入统计:")
        click.echo(f"  新增资产: {len(new_assets)} 条")
        if mode == 'merge-remark':
            click.echo(f"  更新备注: {len(update_assets)} 条")
        else:
            click.echo(f"  跳过重复: {len(db_duplicates)} 条")
        if file_duplicates:
            click.echo(f"  文件内重复(已去重): {len(file_duplicates)} 个编号")

        if dry_run:
            click.echo("\n[试运行] 以上为预览结果，未实际写入数据库")
            return

        imported = 0
        updated = 0
        skipped = 0

        for asset in deduped_list:
            existing = get_asset_by_no(conn, asset['asset_no'])
            if existing:
                if mode == 'merge-remark':
                    old_remark = existing['remark'] or ''
                    new_remark = asset.get('remark', '') or ''
                    merged_remark = old_remark
                    if new_remark:
                        if old_remark:
                            merged_remark = old_remark + '; ' + new_remark
                        else:
                            merged_remark = new_remark

                    update_data = {'remark': merged_remark}
                    for key in ['department', 'location', 'user_name']:
                        if asset.get(key) and not existing.get(key):
                            update_data[key] = asset[key]

                    if update_data:
                        update_asset(conn, asset['asset_no'], update_data)
                        update_asset_timestamp(conn, asset['asset_no'])
                        log_operation(
                            conn,
                            asset['asset_no'],
                            '导入更新',
                            operator,
                            f"合并备注及补全字段: {', '.join(update_data.keys())}"
                        )
                        updated += 1
                    else:
                        skipped += 1
                else:
                    skipped += 1
                continue

            insert_asset(conn, asset)
            log_operation(
                conn,
                asset['asset_no'],
                '导入',
                operator,
                f"导入资产: {asset.get('name', '')}"
            )
            imported += 1

    click.echo(f"\n导入完成: 新增 {imported} 条，更新 {updated} 条，跳过 {skipped} 条")


@click.command('check-duplicates')
@click.option('--file', '-f', 'filepath', type=click.Path(exists=True),
              help='检查指定文件中的重复编号')
@click.option('--db', is_flag=True, help='检查数据库中的重复编号（默认）')
@click.option('--output', '-o', help='导出重复清单到文件')
def check_duplicates_cmd(filepath, db, output):
    """检查重复编号"""
    if filepath:
        ext = os.path.splitext(filepath)[1].lower()
        if ext == '.csv':
            raw_assets = _read_csv(filepath)
        elif ext in ['.xlsx', '.xls']:
            raw_assets = _read_excel(filepath)
        else:
            click.echo(f"错误: 不支持的文件格式 {ext}")
            return

        normalized_list = []
        for raw in raw_assets:
            norm = _normalize_asset(raw)
            if norm:
                normalized_list.append(norm)

        file_dups = _find_file_duplicates(normalized_list)

        if not file_dups:
            click.echo("文件中没有重复编号")
        else:
            click.echo(f"文件中发现 {len(file_dups)} 个重复编号:")
            table = [[no, cnt] for no, cnt in sorted(file_dups.items())]
            click.echo(tabulate(table, headers=['资产编号', '出现次数'], tablefmt='simple'))

            if output:
                with open(output, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['资产编号', '出现次数'])
                    writer.writerows(table)
                click.echo(f"\n重复清单已导出到: {output}")

        with get_db() as conn:
            db_dups = []
            seen = set()
            for asset in normalized_list:
                no = asset['asset_no']
                if no in seen:
                    continue
                seen.add(no)
                if asset_exists(conn, no):
                    db_dups.append(no)

            if db_dups:
                click.echo(f"\n其中 {len(db_dups)} 个已存在于数据库:")
                for no in db_dups[:10]:
                    click.echo(f"  {no}")
                if len(db_dups) > 10:
                    click.echo(f"  ... 还有 {len(db_dups) - 10} 个")
            else:
                click.echo("\n这些编号在数据库中均不存在")
    else:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT asset_no, COUNT(*) as cnt
                FROM assets
                GROUP BY asset_no
                HAVING cnt > 1
            ''')
            dupes = cursor.fetchall()

        if not dupes:
            click.echo("数据库中没有重复编号")
            return

        click.echo(f"数据库中发现 {len(dupes)} 个重复编号:")
        table = [[row['asset_no'], row['cnt']] for row in dupes]
        click.echo(tabulate(table, headers=['资产编号', '出现次数'], tablefmt='simple'))

        if output:
            with open(output, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['资产编号', '出现次数'])
                writer.writerows(table)
            click.echo(f"\n重复清单已导出到: {output}")
