import os
import csv
import re
from datetime import datetime, timedelta
import click
from tabulate import tabulate
from ..database import get_db


@click.command('audit')
@click.option('--repair-days', type=int, default=30, help='维修超期天数阈值（默认30天）')
@click.option('--idle-days', type=int, default=90, help='长期闲置天数阈值（默认90天）')
@click.option('--type', '-t', 'audit_types',
              type=click.Choice(['overdue-repair', 'long-idle', 'dirty-data', 'all']),
              multiple=True, help='审计类型，可多选')
@click.option('--department', '-d', help='按部门筛选')
@click.option('--category', '-c', help='按类别筛选')
@click.option('--export', '-e', help='导出审计报告 (CSV/Excel)')
@click.option('--export-todo', help='导出待处理清单 (CSV，可直接用于 batch)')
@click.option('--format', '-f', 'fmt', type=click.Choice(['csv', 'xlsx']),
              default='csv', help='导出格式')
def audit_cmd(repair_days, idle_days, audit_types, department, category, export, export_todo, fmt):
    """审计异常资产（超期维修、长期闲置、脏数据等）"""
    if not audit_types:
        audit_types = ['all']

    do_all = 'all' in audit_types

    with get_db() as conn:
        results = {}

        if do_all or 'overdue-repair' in audit_types:
            results['overdue_repair'] = _audit_overdue_repair(conn, repair_days, department, category)

        if do_all or 'long-idle' in audit_types:
            results['long_idle'] = _audit_long_idle(conn, idle_days, department, category)

        if do_all or 'dirty-data' in audit_types:
            results['dirty_data'] = _audit_dirty_data(conn, department, category)

    total = sum(len(v) for v in results.values())

    click.echo("=" * 90)
    click.echo("  资产审计报告")
    click.echo("=" * 90)
    click.echo(f"  阈值: 维修超期 > {repair_days} 天，长期闲置 > {idle_days} 天")
    if department:
        click.echo(f"  部门: {department}")
    if category:
        click.echo(f"  类别: {category}")
    click.echo(f"  共发现 {total} 项异常")
    click.echo("=" * 90)
    click.echo()

    if 'overdue_repair' in results:
        _print_overdue_repair(results['overdue_repair'])

    if 'long_idle' in results:
        _print_long_idle(results['long_idle'])

    if 'dirty_data' in results:
        _print_dirty_data(results['dirty_data'])

    if export:
        _export_audit_results(export, fmt, results, repair_days, idle_days, department, category)
        click.echo(f"\n审计报告已导出到: {export}")

    if export_todo:
        _export_todo_list(export_todo, results)
        click.echo(f"\n待处理清单已导出到: {export_todo}")


def _audit_overdue_repair(conn, days, department, category):
    """审计超期维修中的资产"""
    cursor = conn.cursor()
    threshold_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')

    query = '''
        SELECT a.asset_no, a.name, a.category, a.department, a.status, a.user_name,
               ol.created_at as repair_start, ol.detail as repair_detail
        FROM assets a
        JOIN operation_logs ol ON a.asset_no = ol.asset_no
        WHERE a.status = '维修中'
          AND ol.operation = '维修发起'
          AND ol.created_at <= ?
          AND NOT EXISTS (
              SELECT 1 FROM operation_logs ol2
              WHERE ol2.asset_no = a.asset_no
                AND ol2.operation = '维修完成'
                AND ol2.created_at > ol.created_at
          )
    '''
    params = [threshold_date]

    if department:
        query += ' AND a.department = ?'
        params.append(department)
    if category:
        query += ' AND a.category = ?'
        params.append(category)

    query += ' ORDER BY ol.created_at ASC'

    cursor.execute(query, params)
    rows = cursor.fetchall()

    result = []
    for row in rows:
        start_str = row['repair_start']
        try:
            start_dt = datetime.strptime(start_str, '%Y-%m-%d %H:%M:%S')
            repair_days_val = (datetime.now() - start_dt).days
        except (ValueError, TypeError):
            repair_days_val = 0

        detail = row['repair_detail'] or ''
        desc_match = re.search(r'故障[:：]\s*(.+?)(?:\s*\||$)', detail)
        description = desc_match.group(1).strip() if desc_match else ''

        result.append({
            'type': '超期维修',
            'asset_no': row['asset_no'],
            'name': row['name'],
            'category': row['category'],
            'department': row['department'] or '',
            'status': row['status'],
            'user_name': row['user_name'] or '',
            'repair_days': repair_days_val,
            'repair_start': start_str,
            'description': description,
            'suggested_action': '补录维修完成记录，或跟进维修进度',
            'suggested_batch_op': 'remark',
            'suggested_value': '维修超期，请跟进',
        })

    return result


def _audit_long_idle(conn, days, department, category):
    """审计长期闲置的资产（闲置状态且无使用人）"""
    cursor = conn.cursor()
    threshold_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')

    query = '''
        SELECT * FROM assets
        WHERE status = '闲置'
          AND (user_name IS NULL OR user_name = '')
          AND updated_at <= ?
    '''
    params = [threshold_date]

    if department:
        query += ' AND department = ?'
        params.append(department)
    if category:
        query += ' AND category = ?'
        params.append(category)

    query += ' ORDER BY updated_at ASC'

    cursor.execute(query, params)
    rows = cursor.fetchall()

    result = []
    for row in rows:
        updated_str = row['updated_at']
        try:
            updated_dt = datetime.strptime(updated_str, '%Y-%m-%d %H:%M:%S')
            idle_days_val = (datetime.now() - updated_dt).days
        except (ValueError, TypeError):
            idle_days_val = 0

        result.append({
            'type': '长期闲置',
            'asset_no': row['asset_no'],
            'name': row['name'],
            'category': row['category'],
            'department': row['department'] or '',
            'status': row['status'],
            'user_name': '',
            'idle_days': idle_days_val,
            'last_user': row['user_name'] or '',
            'last_change': updated_str,
            'suggested_action': '建议分配给有需要的员工，或评估后报废',
            'suggested_batch_op': 'remark',
            'suggested_value': '长期闲置，请安排处理',
        })

    return result


def _audit_dirty_data(conn, department, category):
    """审计脏数据，按子类型分类"""
    cursor = conn.cursor()
    issues = []

    query1 = '''
        SELECT * FROM assets
        WHERE status = '已报废' AND user_name IS NOT NULL AND user_name != ''
    '''
    params = []
    if department:
        query1 += ' AND department = ?'
        params.append(department)
    if category:
        query1 += ' AND category = ?'
        params.append(category)

    cursor.execute(query1, params)
    for row in cursor.fetchall():
        issues.append({
            'type': '脏数据-已报废有使用人',
            'sub_type': '已报废有使用人',
            'asset_no': row['asset_no'],
            'name': row['name'],
            'category': row['category'],
            'department': row['department'] or '',
            'status': row['status'],
            'user_name': row['user_name'] or '',
            'issue': '已报废但仍有使用人',
            'suggested_action': '清空使用人，确保报废资产无责任人',
            'suggested_batch_op': 'remark',
            'suggested_value': '已清空使用人（脏数据修复）',
        })

    query2 = '''
        SELECT * FROM assets
        WHERE status = '在用' AND (user_name IS NULL OR user_name = '')
    '''
    params2 = []
    if department:
        query2 += ' AND department = ?'
        params2.append(department)
    if category:
        query2 += ' AND category = ?'
        params2.append(category)

    cursor.execute(query2, params2)
    for row in cursor.fetchall():
        issues.append({
            'type': '脏数据-在用无使用人',
            'sub_type': '在用无使用人',
            'asset_no': row['asset_no'],
            'name': row['name'],
            'category': row['category'],
            'department': row['department'] or '',
            'status': row['status'],
            'user_name': '',
            'issue': '在用状态但无使用人',
            'suggested_action': '补录分配记录，或标记为闲置',
            'suggested_batch_op': 'remark',
            'suggested_value': '在用无使用人，请核实',
        })

    issues.sort(key=lambda x: (x['sub_type'], x['asset_no']))
    return issues


def _print_overdue_repair(items):
    click.echo(f"▶ 超期维修中资产（{len(items)} 项）")
    click.echo("-" * 90)
    if items:
        rows = []
        for item in items:
            rows.append([
                item['asset_no'], item['name'], item['department'],
                item['repair_days'], item['repair_start'],
                item['description'] or '-', item['suggested_action'],
            ])
        click.echo(tabulate(rows,
                            headers=['资产编号', '资产名称', '部门', '维修天数', '开始时间',
                                     '故障描述', '建议动作'],
                            tablefmt='simple'))
    else:
        click.echo("（无异常）")
    click.echo()


def _print_long_idle(items):
    click.echo(f"▶ 长期闲置资产（{len(items)} 项）")
    click.echo("-" * 90)
    if items:
        rows = []
        for item in items:
            rows.append([
                item['asset_no'], item['name'], item['department'],
                item['idle_days'], item['last_user'] or '-',
                item['last_change'], item['suggested_action'],
            ])
        click.echo(tabulate(rows,
                            headers=['资产编号', '资产名称', '部门', '闲置天数',
                                     '最后使用人', '最后变更时间', '建议动作'],
                            tablefmt='simple'))
    else:
        click.echo("（无异常）")
    click.echo()


def _print_dirty_data(items):
    sub_types = {}
    for item in items:
        st = item.get('sub_type', '其他')
        if st not in sub_types:
            sub_types[st] = []
        sub_types[st].append(item)

    click.echo(f"▶ 脏数据（{len(items)} 项）")
    click.echo("-" * 90)

    if not items:
        click.echo("（无异常）")
        click.echo()
        return

    for st, sub_items in sorted(sub_types.items()):
        click.echo(f"  · {st}（{len(sub_items)} 项）")
        if sub_items:
            rows = []
            for item in sub_items:
                rows.append([
                    item['asset_no'], item['name'], item['department'],
                    item['status'], item['user_name'] or '-',
                    item['suggested_action'],
                ])
            click.echo(tabulate(rows,
                                headers=['资产编号', '资产名称', '部门', '状态',
                                         '使用人', '建议动作'],
                                tablefmt='simple'))
        click.echo()


def _export_audit_results(filepath, fmt, results, repair_days, idle_days, department, category):
    """导出审计报告"""
    ext = os.path.splitext(filepath)[1].lower()

    if fmt == 'xlsx' or ext in ['.xlsx', '.xls']:
        try:
            from openpyxl import Workbook
            wb = Workbook()

            ws = wb.active
            ws.title = '汇总'
            ws.append(['资产审计报告'])
            ws.append(['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            ws.append(['维修超期阈值(天)', repair_days])
            ws.append(['长期闲置阈值(天)', idle_days])
            if department:
                ws.append(['部门', department])
            if category:
                ws.append(['类别', category])
            ws.append([])
            ws.append(['分类', '数量', '建议动作'])
            for section_name, items in [
                ('超期维修中资产', results.get('overdue_repair', [])),
                ('长期闲置资产', results.get('long_idle', [])),
                ('脏数据', results.get('dirty_data', [])),
            ]:
                sample_action = items[0]['suggested_action'] if items else ''
                ws.append([section_name, len(items), sample_action])

            if 'overdue_repair' in results:
                ws2 = wb.create_sheet('超期维修')
                ws2.append(['资产编号', '资产名称', '类别', '部门', '状态', '使用人',
                            '维修天数', '开始时间', '故障描述', '建议动作'])
                for item in results['overdue_repair']:
                    ws2.append([
                        item['asset_no'], item['name'], item['category'],
                        item['department'], item['status'], item['user_name'],
                        item['repair_days'], item['repair_start'],
                        item['description'], item['suggested_action'],
                    ])

            if 'long_idle' in results:
                ws3 = wb.create_sheet('长期闲置')
                ws3.append(['资产编号', '资产名称', '类别', '部门',
                            '闲置天数', '最后使用人', '最后变更时间', '建议动作'])
                for item in results['long_idle']:
                    ws3.append([
                        item['asset_no'], item['name'], item['category'],
                        item['department'], item['idle_days'],
                        item['last_user'], item['last_change'],
                        item['suggested_action'],
                    ])

            if 'dirty_data' in results:
                ws4 = wb.create_sheet('脏数据')
                ws4.append(['资产编号', '资产名称', '类别', '部门',
                            '状态', '使用人', '问题类型', '建议动作'])
                for item in results['dirty_data']:
                    ws4.append([
                        item['asset_no'], item['name'], item['category'],
                        item['department'], item['status'],
                        item['user_name'], item.get('sub_type', item['type']),
                        item['suggested_action'],
                    ])

            wb.save(filepath)
        except ImportError:
            click.echo("错误: 需要安装 openpyxl 才能导出 Excel 文件")
    else:
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['资产审计报告'])
            writer.writerow(['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['维修超期阈值(天)', repair_days])
            writer.writerow(['长期闲置阈值(天)', idle_days])
            if department:
                writer.writerow(['部门', department])
            if category:
                writer.writerow(['类别', category])
            writer.writerow([])

            if 'overdue_repair' in results:
                items = results['overdue_repair']
                writer.writerow([f'=== 超期维修中资产 ({len(items)} 项) ==='])
                writer.writerow(['资产编号', '资产名称', '类别', '部门', '状态', '使用人',
                                 '维修天数', '开始时间', '故障描述', '建议动作'])
                for item in items:
                    writer.writerow([
                        item['asset_no'], item['name'], item['category'],
                        item['department'], item['status'], item['user_name'],
                        item['repair_days'], item['repair_start'],
                        item['description'], item['suggested_action'],
                    ])
                writer.writerow([])

            if 'long_idle' in results:
                items = results['long_idle']
                writer.writerow([f'=== 长期闲置资产 ({len(items)} 项) ==='])
                writer.writerow(['资产编号', '资产名称', '类别', '部门',
                                 '闲置天数', '最后使用人', '最后变更时间', '建议动作'])
                for item in items:
                    writer.writerow([
                        item['asset_no'], item['name'], item['category'],
                        item['department'], item['idle_days'],
                        item['last_user'], item['last_change'],
                        item['suggested_action'],
                    ])
                writer.writerow([])

            if 'dirty_data' in results:
                items = results['dirty_data']
                writer.writerow([f'=== 脏数据 ({len(items)} 项) ==='])
                writer.writerow(['资产编号', '资产名称', '类别', '部门',
                                 '状态', '使用人', '问题类型', '建议动作'])
                for item in items:
                    writer.writerow([
                        item['asset_no'], item['name'], item['category'],
                        item['department'], item['status'],
                        item['user_name'], item.get('sub_type', item['type']),
                        item['suggested_action'],
                    ])
                writer.writerow([])


def _export_todo_list(filepath, results):
    """导出待处理清单（可直接用于 batch 批量处理）"""
    todo_items = []

    for section, items in results.items():
        for item in items:
            todo_items.append({
                '异常类型': item['type'],
                '资产编号': item['asset_no'],
                '资产名称': item['name'],
                '类别': item['category'],
                '部门': item['department'],
                '当前状态': item['status'],
                '使用人': item.get('user_name', ''),
                '问题描述': item.get('description') or item.get('issue') or item.get('type', ''),
                '建议动作': item['suggested_action'],
                '建议批量操作': item.get('suggested_batch_op', 'remark'),
                '建议备注': item.get('suggested_value', ''),
            })

    if not todo_items:
        click.echo("\n没有待处理项，跳过导出")
        return

    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=todo_items[0].keys())
        writer.writeheader()
        writer.writerows(todo_items)
